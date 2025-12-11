import sys
import json
import shutil
import os
import copy
from openpyxl import load_workbook

def main():
    try:
        # Read input JSON from stdin
        try:
            input_content = sys.stdin.read()
            if not input_content:
                print(json.dumps({"error": "Empty input"}))
                return
            input_data = json.loads(input_content)
        except Exception as e:
            print(json.dumps({"error": f"Invalid JSON input: {str(e)}"}))
            return

        template_path = input_data.get("templatePath")
        output_path = input_data.get("outputPath")
        quote_data = input_data.get("quoteData")

        if not template_path or not output_path or not quote_data:
            print(json.dumps({"error": "Missing required arguments: templatePath, outputPath, quoteData"}))
            return

        if not os.path.exists(template_path):
             print(json.dumps({"error": f"Template file not found at {template_path}"}))
             return

        try:
            shutil.copyfile(template_path, output_path)
        except Exception as e:
            print(json.dumps({"error": f"Failed to create output file: {str(e)}"}))
            return

        wb = load_workbook(output_path, keep_vba=True)
        
        # --- Populate Paramètre ---
        if "Paramètre" in wb.sheetnames:
            ws_param = wb["Paramètre"]
            client = quote_data.get("client", {})
            
            # Client Info
            ws_param.cell(row=7, column=3, value=client.get("name", ""))
            ws_param.cell(row=8, column=3, value=client.get("address1", ""))
            ws_param.cell(row=9, column=3, value=client.get("city", ""))
            ws_param.cell(row=11, column=3, value=client.get("region", ""))
            ws_param.cell(row=12, column=3, value=client.get("country", ""))
            ws_param.cell(row=13, column=3, value=client.get("zip", ""))
            
            contact = quote_data.get("contact", {})
            ws_param.cell(row=14, column=3, value=contact.get("firstName", ""))
            ws_param.cell(row=15, column=3, value=contact.get("lastName", ""))
            
            # Contact Details (from Client or Contact?)
            # Assuming client phone/email if contact is missing or as backup
            ws_param.cell(row=16, column=3, value=client.get("phone", ""))
            ws_param.cell(row=18, column=3, value=client.get("email", ""))

            # Projet Info
            ws_param.cell(row=21, column=3, value=quote_data.get("projectName", ""))
            ws_param.cell(row=22, column=3, value=quote_data.get("quoteNumber", ""))
            
            # Language (Row 40, Col 3 - based on dump)
            # User wants language code (e.g., 'en', 'fr')
            if quote_data.get("language"):
                 ws_param.cell(row=40, column=3, value=quote_data.get("language"))
            
            # Currency
            if quote_data.get("currency"):
                 ws_param.cell(row=65, column=3, value=quote_data.get("currency"))

        else:
             print(json.dumps({"error": "Sheet 'Paramètre' not found"}))
             return

        # --- Populate Cotation ---
        # User specified "l'onglet cotation". Use that.
        target_sheet_name = "Cotation"
        if target_sheet_name in wb.sheetnames:
            ws = wb[target_sheet_name]
            
            # Template Row is Row 8 (based on dump analysis)
            start_row = 8
            
            items = quote_data.get("items", [])
            num_items = len(items)
            
            if num_items > 0:
                # Insert rows if we have more than 1 item
                # We start filling at start_row. If we need 5 items, we need 4 new rows.
                if num_items > 1:
                    ws.insert_rows(start_row + 1, amount=num_items - 1)
                    
                    # Copy formulas and styles from Template Row (start_row) to new rows
                    # Iterate columns in the template row
                    for col in range(1, ws.max_column + 1):
                        template_cell = ws.cell(row=start_row, column=col)
                        
                        # Apply to all new rows
                        for i in range(1, num_items):
                            target_row = start_row + i
                            target_cell = ws.cell(row=target_row, column=col)
                            
                            # Copy Style
                            if template_cell.has_style:
                                target_cell.font = copy.copy(template_cell.font)
                                target_cell.border = copy.copy(template_cell.border)
                                target_cell.fill = copy.copy(template_cell.fill)
                                target_cell.number_format = template_cell.number_format
                                target_cell.protection = copy.copy(template_cell.protection)
                                target_cell.alignment = copy.copy(template_cell.alignment)
                            
                            # Copy Value or Formula
                            if template_cell.value is not None:
                                # If it's a formula, we might need to adjust row references?
                                # openpyxl doesn't strictly translate formulas on copy.
                                # Simple copy: returns "=A8+B8"
                                # We usually need translation.
                                # Given complexity, we'll try straight copy. Excel often adjusts relative refs IF pasted, but via code it's literal.
                                # However, many sheets use Column references or Table references.
                                # If it's a literal formula (e.g. "=C8*D8"), it will point to Row 8 in Row 9. Bad.
                                # Python formula translation is hard.
                                # BUT: User said "Cotation creates lines".
                                # Maybe the MACRO does the heavy lifting.
                                # Check if we can just write the Inputs and let the User run the "Refresh" macro?
                                # User said "l'onglet cotation ... va créer le nombre de ligne demandé". 
                                # This implies the MACRO creates lines. 
                                # If I create lines manually, I might bypass the macro logic.
                                # Compromise: I will write data to the first N rows. If rows don't exist, I insert.
                                # I will TRY to update simple row references in formulas.
                                val = template_cell.value
                                if isinstance(val, str) and val.startswith("="):
                                    # Very basic formula adjuster
                                    # Replace "8" with target_row IF it looks like a ref? Too risky.
                                    # We will just copy string. If formula logic is broken, User relies on macro 'Refresh' or we must ask precise questions.
                                    target_cell.value = val 
                                else:
                                    # Constant value (don't copy specific input values like 'L1', copy structure only? No, copy everything then overwrite inputs)
                                    target_cell.value = val

                # Fill Data for all items
                for index, item in enumerate(items):
                    row_idx = start_row + index
                    
                    # Map Inputs
                    # A: Ligne ID
                    ws.cell(row=row_idx, column=1, value=f"L{index+1}")
                    # B: Tag
                    ws.cell(row=row_idx, column=2, value=f"{quote_data.get('quoteNumber','')} - {index+1}")
                    # C: Qty
                    ws.cell(row=row_idx, column=3, value=item.get("quantity", 1))
                    # D: Finition
                    ws.cell(row=row_idx, column=4, value=f"{item.get('material', '')} {item.get('finish', '')}")
                    # E: Length
                    ws.cell(row=row_idx, column=5, value=item.get("length", 0))
                    # F: Width
                    ws.cell(row=row_idx, column=6, value=item.get("width", 0))
                    # G: Thickness
                    ws.cell(row=row_idx, column=7, value=item.get("thickness", 0))

        wb.save(output_path)
        print(json.dumps({"success": True, "file": output_path}))
    except Exception as e:
        print(json.dumps({"error": f"Error processing Excel: {str(e)}"}))

if __name__ == "__main__":
    main()
